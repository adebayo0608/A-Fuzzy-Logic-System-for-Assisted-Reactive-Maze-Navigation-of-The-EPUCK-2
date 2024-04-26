# THIS FILE CONTAINS ALL FUNCTIONS FOR THE OBSTACLE AVOIDANCE OF THE EPUCK 2
# USES A PS4 CONTROLLER

import os
import pygame
from pygame._sdl2 import controller as controller_module
from tabulate import tabulate 
import numpy as np
from typing import List
import pandas as pd
import time
import openpyxl
from openpyxl.utils import get_column_letter

# Constants for the states
USER_CONTROLLING = "USER_CONTROLLING"
AVOIDING_OBSTACLE = "AVOIDING_OBSTACLE"
STATIC_OPERATOR_REQUIRED = "STATIC_OPERATOR_REQUIRED"

# This class implements control of the EPUCK with variable speed 
class fundamentalCommandExecution:
	
	# Constructor
	def __init__(self, controller_index=0,):
		# Initialize pygame and the controller module
		pygame.init()
		controller_module.init()

		# Verify that 1 or more controllers are connected
		if controller_module.get_count() > 0:
			# Initialising the specified PS4 controller
			self.ps4_controller = controller_module.Controller(controller_index)
		else:
			raise Exception("CONTROLLER IS NOT CONNECTED.")

		# Initialize class booleans
		
		# Booleans regarding the joystick
		self.is_l_joystick_moved = False
		self.is_r_joystick_moved = False
		self.is_moving_forward = False
		self.is_moving_reverse = False
		self.is_rotating_left = False
		self.is_rotating_right = False     
		
		# Booleans regarding safety
		self.is_forward_allowed = True
		self.is_reverse_allowed = True
		self.is_left_rotation_allowed = True
		self.is_right_rotation_allowed = True   
		
		# Initialising class variables
		self.max_speed_rpm = 447               # 1000 [rpm]                
		self.rotation_speed_factor = 0.1       # Rotation sensitivity 
		self.motor_speed = [0,0]           
		self.safety_distance = 0.056   
		self.proximity_sensors = [0] * 16  
		self.actual_proximity_distances = [] 
		self.forward_distance = 0
		self.actual_forward_distance = 0
		   
	# GETTERS - To be used my main code
	def get_speed(self): 
		return self.motor_speed     
   
	def set_led_colour(self, colour):
		"""Set the LED color.
		Args:
			colour (str): The color to set. Valid values: 'red', 'green', 'orange', 'yellow', 'purple'.

		Raises:
			ValueError: If an invalid color is specified.
		"""
		colour_lower = colour.lower()
		
		if colour_lower == 'red':
			self.led_array = [100, 0, 0, 100, 0, 0, 100, 0, 0, 100, 0, 0]
		elif colour_lower == 'green':
			self.led_array = [0, 100, 0, 0, 100, 0, 0, 100, 0, 0, 100, 0]
		elif colour_lower == 'orange':
			self.led_array = [100, 50, 0, 100, 50, 0, 100, 50, 0, 100, 50, 0]
		elif colour_lower == 'yellow':
			self.led_array = [100, 100, 0, 100, 100, 0, 100, 100, 0, 100, 100, 0]
		elif colour_lower == 'purple':
			self.led_array = [50, 0, 50, 50, 0, 50, 50, 0, 50, 50, 0, 50]
		else:
			raise ValueError("Invalid colour specified. Available colours: 'red', 'green', 'orange', 'yellow', 'purple'")   
	
	def get_led_colour(self):
		"""
		Returns:
			Array: 12 elements representing the LED Colour.
		"""
		return self.led_array
	
	def reset_inputs(self):
		self.is_l_joystick_moved = False
		self.is_r_joystick_moved = False
		self.is_moving_forward = False
		self.is_moving_reverse = False
		self.is_rotating_left = False
		self.is_rotating_right = False
		
	def calibrate_joystick(self, joystick):
		"""Eliminates Stick Drift of the input joystick value"""
		
		if joystick > -3000 and joystick < 3000 :
			answer = 0       
		else:
			answer = joystick
		return answer            
			
	def check_left_joystick(self):
		"""Checks if the left joystick has been moved"""
		left_joystick_y = self.ps4_controller.get_axis(pygame.CONTROLLER_AXIS_LEFTY)
		left_joystick_y = self.calibrate_joystick(left_joystick_y)

		if left_joystick_y < 0:
			self.is_l_joystick_moved = True
			self.is_moving_forward = True
			
		elif left_joystick_y > 0:
			self.is_l_joystick_moved = True
			self.is_moving_reverse = True
			
		else:
			self.is_l_joystick_moved = False
 
	def check_right_joystick(self):
		"""Checks if the right joystick has been moved"""
		right_joystick_x = self.ps4_controller.get_axis(pygame.CONTROLLER_AXIS_RIGHTX)
		right_joystick_x = self.calibrate_joystick(right_joystick_x)

		if right_joystick_x < 0:
			self.is_r_joystick_moved = True 
			self.is_rotating_left = True
			
		elif right_joystick_x > 0:
			self.is_r_joystick_moved = True
			self.is_rotating_right = True            
			
		else:
			self.is_r_joystick_moved = False

	def check_exit_button(self):
		"""Checks if the exit button has been pressed
		"""
		exit_button_pressed = self.ps4_controller.get_button(pygame.CONTROLLER_BUTTON_B)
		if exit_button_pressed:
			print(" 'O' Pressed.")
			print("EPUCK 2 Session Terminated.")
			
	def check_inputs(self):
		"""
		Checks the states of the joysticks and buttons"""
		self.reset_inputs()
		for event in pygame.event.get():
			if event.type == pygame.JOYAXISMOTION:
				self.check_left_joystick()
				self.check_right_joystick()
			elif event.type == pygame.JOYBUTTONDOWN:
				self.check_exit_button()
	
	def quit(self):
		"""Quits the application and cleans up resources.
		"""        
		controller_module.quit()
		pygame.quit()     
		
	def store_sensor_readings(self,proximity_sensors: List[int], forward_distance):
		"""
		Stores sensor data.

		Args:
			proximity_sensors (list of int): List of proximity sensor readings.
			forward_distance (float): Forward distance reading in cm.
		"""
		self.proximity_sensors = proximity_sensors
		self.forward_distance = forward_distance    
	
	def calculate_actual_distance(self,proximity_reading):
		"""
		Calculate the actual distance corresponding to a given proximity reading.
		
		Args:
			proximity_reading (float): Proximity reading value.
		
		Returns:
			float: Actual distance.
		"""
		
		# Fixed optimal parameters obtained from curve fitting
		a = 1521.7170108943192
		b = -1.256270043608874
		c = -118.51612931657212
			
		# Calculate actual distance using the inverse of the fitted function
		actual_distance_cm = ((proximity_reading - c) / a) ** (1 / b)
		
		actual_distance_m = round(actual_distance_cm / 100 , 3)     
		
		return actual_distance_m 
	
	def calculate_actual_forward_distance(self, forward_distance_cm):
		"""
		Calculate the actual distance based on the linear relationship between measured and actual distance.

		Args:
			reading (float): The reading obtained from the sensor [cm].

		Returns:
			float: The actual distance calculated from the reading.
		"""
		# Fixed slope and intercept : y = mx + c
		slope = 1.001129343                            #0.997
		intercept = -2.726898293                       #-2.686

		actual_distance = slope * forward_distance_cm + intercept
		return round(actual_distance/100, 3)
	
	def convert_sensor_data_to_actual_distances(self):
		"""
		Convert filtered sensor data to actual distances.
		"""
		# Convert proximity sensor readings to actual distances
		self.actual_proximity_distances = [self.calculate_actual_distance(sensor_readings) for sensor_readings in self.proximity_sensors]
		
		# Convert forward distance reading to actual distance
		self.actual_forward_distance = self.calculate_actual_forward_distance(self.forward_distance) 
	
	def process_sensors_user_mode(self,proximity_sensors: List[int], forward_distance):
		"""
		Store sensor data and convert them to actual distances

		Args:
			proximity_sensors (List[int]): List of proximity sensor readings
			forward_distance (float): Forward distance [cm]
		"""
		self.store_sensor_readings(proximity_sensors, forward_distance)
		self.convert_sensor_data_to_actual_distances()
	
	def safety_check(self): 
		"""
		Ensures safe movement by checking proximity sensor readings and forward distance.

		This function restricts forward or reverse movement if the robot is too close to obstacles.        
		Direction of restriction depends on direction of the obstacle.
		"""                     
		# Sensors at the front of the EPUCK       
		if self.actual_proximity_distances[0] <= self.safety_distance or self.actual_proximity_distances[7] <= self.safety_distance:
			self.is_forward_allowed = False
		# Sensors at the back of the EPUCK
		if self.actual_proximity_distances[3] <= self.safety_distance or self.actual_proximity_distances[4] <= self.safety_distance:  
			self.is_reverse_allowed = False

	def adjust_speed_for_safety(self, speed):
		"""
		Adjust the motor speed based on safety restrictions.

		Args:
			speed (float): The calculated motor speed.

		Returns:
			float: The adjusted motor speed.
		"""
		if not self.is_forward_allowed and self.is_moving_forward:
			# If object in forward direction and user wants to move forward 
			speed = 0 
				
		if not self.is_reverse_allowed and self.is_moving_reverse:
			# If object in backward direction and user wants to reverse
			speed = 0             
		
		return speed  
	
	def determine_speed(self, speed, rotation_speed):
		"""
		Determine motor speed based on joystick input and safety restrictions.

		Args:
			speed (float): Speed based on joystick input.
			rotation_speed (float): Rotation speed based on joystick input.

		Returns:
			list: A list containing the motor speeds for the left and right motors.
		"""
		
		if self.is_rotating_left and not self.is_l_joystick_moved:
			# Static Left Rotation: Robot is stationary and rotating left
			return [- 3/4 * self.max_speed_rpm,  3/4 *self.max_speed_rpm]
		elif self.is_rotating_left and self.is_l_joystick_moved:
			# Moving Left Rotation: Robot is moving forward while turning left
			return [rotation_speed, speed]
		elif self.is_rotating_right and not self.is_l_joystick_moved:
			# Static Right Rotation: Robot is stationary and rotating right
			return [ 3/4 *self.max_speed_rpm, - 3/4 *self.max_speed_rpm]
		elif self.is_rotating_right and self.is_l_joystick_moved:
			# Moving Right Rotation: Robot is moving forward while turning right
			return [speed, rotation_speed]
		else:
			# Straight Linear Movement: Robot is moving straight forward or backward
			return [speed, speed] 
	 
	def calibrate_speed(self,speed):
		"""_summary_
		"""
		if speed < 30 and speed > -30:
			speed = 0
		
		return speed
			   
	def calculate_speed(self):
		"""
		Calculate motor speed based on joystick input.
		
		Speed is set to 0 to prevent collisions before they occur.
		"""
		# Modfied here: Safety Check
		self.safety_check() 
		print("Safety Check performed")
		
		l_joystick_y_value = -(self.ps4_controller.get_axis(pygame.CONTROLLER_AXIS_LEFTY))
		
		# Calculate speed proportionally
		speed = np.interp(l_joystick_y_value, [-32678, 32767], [-self.max_speed_rpm, self.max_speed_rpm])

		# Ensure the speed is within the range of -max_speed to max_speed
		speed = (max(-self.max_speed_rpm, min(speed, self.max_speed_rpm)))            
		
		# Adjust the speed if required (safety reasons)
		speed = self.adjust_speed_for_safety(speed)  
		
		speed = self.calibrate_speed(speed)          
		
		rotation_speed = speed * self.rotation_speed_factor
		
		# Ensure rotation speed is within the valid range
		rotation_speed = max(-32767, min(rotation_speed, 32767))
		
		# Determine speed based on calculated speed and rotation speed
		self.motor_speed = self.determine_speed(speed,rotation_speed)  
  
	def run_user_controlling(self,proximity_sensor_readings, forward_distance_cm):
		self.process_sensors_user_mode(proximity_sensor_readings, forward_distance_cm)
		self.set_led_colour('green')
		self.check_inputs()
		self.calculate_speed()
	
	# ********************************************************************** TEST FUNCTIONS ****************************************************************************************#        
	def test_safety(self):
		"""
		Test the safety mechanism by printing sensor readings and movement allowance status.
		
		This function displays the current proximity sensor readings and forward distance, 
		along with whether forward and reverse movements are allowed based on the safety mechanism.
		"""
		
		print("TESTING: SAFETY")
		print()
		
		print("Safety Distance: ",self.safety_distance)
		print() 
		
		print("Proximity Readings:")
		print("Front Centre: ", self.actual_forward_distance)
		print("Front Left: ", self.actual_proximity_distances[0])
		print("Front Right: ", self.actual_proximity_distances[7])
		print("Back Left: ", self.actual_proximity_distances[3])
		print("Back Right: ", self.actual_proximity_distances[4]) 
		print()
		
		if self.is_forward_allowed:
			print("Forward Movement: Allowed")
		else:
			print("Forward Movement: NOT Allowed")
			
		if self.is_reverse_allowed:
			print("Reverse Movement: Allowed")
		else:
			print("Reverse Movement: NOT Allowed")
			
		if self.is_left_rotation_allowed:
			print("Left Rotation: Allowed")
		else:
			print("Left Rotation: NOT Allowed")
			
		if self.is_right_rotation_allowed:
			print("Right Rotation: Allowed")
		else:
			print("Right Rotation: NOT Allowed")
			
		
		
class ObjectAvoidance(fundamentalCommandExecution):
		
	def __init__(self):
		
		# Inherit all variables from previous class        
		super().__init__()
		
		# States
		self.curr_state = None
		self.next_state = None
		
		# EPUCK Geometric Parameters
		self.steps_per_revolution = 20
		self.gear_ratio = 50
		self.wheel_diameter_m = 0.041  
		self.wheel_distance_m = 0.053
		
		# Instance Variables 
		self.radius_obj = 0
		self.forward_distance = 0           # Distance from centre sensor to object   
		self.actual_forward_distance = None # Actual distance ^^
	  
		self.l_l = 0                        # both ll and lr set to 0 = no object on robot's trajectory path
		self.l_r = 0
		self.safety_distance = 0.056         # Safety Distance (m)
		self.max_distance_measured = 0.08   # Detection Distance (m)
		self.v_max =  0.022                 # Maximum linear velocity (m/s) - controls how fast it is during OA mode
		self.w_max = 2                      # Maximum rotational velocity (rad/s) - controls how much it turns in OA mode
		self.side = None
		self.d_or = None       
		self.v_u = 0                        # User linear velovity (m/s) - Input from the controller       
		self.w_u = 0                        # User  rotational velocity (rad/s) - Input from the controller
		self.v_oa = 0                       # OA linear velovity (m/s)       
		self.w_oa = 0                       # OA rotational velocity (rad/s)
		self.v = 0                          # Output linear velovity  (m/s)     
		self.w = 0                          # Output rotational velocity (rad/s)
		self.v_lw_mps = 0                   # Velocity of left wheel (m/s)
		self.v_rw_mps = 0                   # Velocity of right wheel (m/s) 
		self.v_lw_sps = 0                   # Velocity of left wheel (steps/s)
		self.v_rw_sps = 0                   # Velocity of right wheel (steps/s)       
		
		# Instance Variables (Testing) 
		self.forward_distance_ama = 0       # Filtered forward distance: Average moving average 
		self.forward_distance_ema = 0       # Filtered forward distance: Exponential moving average
		self.forward_distance_median = 0    # # Filtered forward distance: Median filter
		
		# Instance Arrays
		self.led_array = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
		self.motor_speed = [0, 0]        
		self.proximity_sensors = [0] * 16
		self.unsafe_sensors = [0] * 16 
		self.left_membership = {}
		self.right_membership = {}
		self.near_membership = {}
		self.far_membership = {}
		self.actual_proximity_distances = []
		
		
		# Instance Booleans        
		self.is_forward_collision_imminent = False
		self.is_reverse_collision_imminent = False
		 
	# GETTERS - To be used in the main code
	def get_velocity(self):
		"""
		Returns:
			Array: [Left motor velocity, Right motor velocity]
			Units: steps/s
		"""
		return self.motor_speed    
	
	def get_curr_state(self):
		"""
		Returns:  Current state of the EPUCK
		
		Possible States: "USER_CONTROLLING", "AVOIDING_OBSTACLE", "STATIC_OPERATOR_REQUIRED"
		"""        
		return self.curr_state
	
	def get_next_state(self):
		"""
		Returns:  Current state of the EPUCK
		
		Possible States: "USER_CONTROLLING", "AVOIDING_OBSTACLE", "STATIC_OPERATOR_REQUIRED"
		"""
		return self.next_state
	
	def convert_mps_to_steps_per_second(self, velocity_mps):
		"""
		Convert velocity from meters per second to steps per second for the e-puck2 robot.

		Args:
			velocity_mps (float): Velocity in meters per second.

		Returns:
			float: Velocity in steps per second.
		"""
		# Calculate the circumference 
		wheel_circumference_meters = np.pi * self.wheel_diameter_m

		# Calculate the distance traveled per step 
		step_distance_meters = wheel_circumference_meters / (self.steps_per_revolution * self.gear_ratio)

		# Convertion: m/s -> steps/s
		velocity_steps_per_second = velocity_mps / step_distance_meters
		return velocity_steps_per_second
	 
	def convert_cm_to_m(self,distance_cm):
		"""Converts from centimetre to metre

		Args:
			distance_cm (float): Distance in centimetre

		Returns:
			float: Distance in metre
		"""
		distance_m = distance_cm / 100
		return distance_m
			
	def convert_forward_distance(self, forward_distance_cm):
		"""Converts the raw forward distance reading to actual distance in meters.

		Args:
			forward_distance_cm (float): Raw Forward Distance (cm).
			
		Returns:
			Float: Actual distance (m)
		"""
		# Constants for linear approximation
		m = -0.0845  # Slope
		b = 12.46    # Intercept

		# Perform linear approximation
		actual_distance_cm = m * forward_distance_cm + b
		
		actual_distance_m = self.convert_cm_to_m(actual_distance_cm)
		
		# Convert: cm -> m
		return actual_distance_m 
			
	def set_velocity(self, left_motor_velocity_mps, right_motor_velocity_mps):
		"""Sets the speed of each wheel
		
		Arguments:
			Units: m/s
		"""
		left_motor_velocity_sps = self.convert_mps_to_steps_per_second(left_motor_velocity_mps)
		righ_motor_velocity_sps = self.convert_mps_to_steps_per_second(right_motor_velocity_mps)
		self.motor_speed = [left_motor_velocity_sps, righ_motor_velocity_sps]
	
	def moving_average_filter(self,proximity_readings):
		"""Uses a moving average filter to calculate the average proximity readings and the average forward distance reading.

		Args:
			proximity_readings (list): A list containing readings for each proximity sensor.
									Each sensor's readings are stored as a list.
			forward_distance_readings (list): A list containing multiple single-element readings
											for the forward distance sensor.

		Returns:
			tuple: A tuple containing two elements:
				- A list of average readings for each proximity sensor.
				- The average reading for the forward distance sensor.
		"""
		# Calculate the average proximity reading for each sensor
		proximity_avg = []
		for sensor_readings in proximity_readings:
			# Calculate the sum of readings for the current sensor
			sensor_sum = sum(sensor_readings)
			# Average reading for the current sensor
			sensor_avg = sensor_sum / len(sensor_readings)
			proximity_avg.append(sensor_avg)
		
		return proximity_avg

	def exponential_moving_average_filter(self,proximity_readings):
		"""Calculates the exponential moving average proximity readings and the exponential moving average forward distance reading.

		Args:
			proximity_readings (list): A list containing readings for each proximity sensor.
									Each sensor's readings are stored as a list.
			forward_distance_readings (list): A list containing multiple single-element readings
											for the forward distance sensor.

		Returns:
			tuple: A tuple containing two elements:
				- A list of exponential moving average readings for each proximity sensor.
				- The exponential moving average reading for the forward distance sensor.
		"""
		alpha = 0.2  # smoothing factor
		
		# Calculate the exponential moving average proximity reading for each sensor
		proximity_ema = []
		for sensor_readings in proximity_readings:
			ema = sensor_readings[0]  # Initialize EMA with the first reading
			for reading in sensor_readings[1:]:
				ema = alpha * reading + (1 - alpha) * ema  # Calculate EMA recursively
			proximity_ema.append(ema)       
	   
		return proximity_ema

	def median_filter(self,proximity_readings):
		"""Applies a median filter to the proximity sensor readings and calculates the median value
		for each sensor, as well as the median value for the forward distance sensor.

		Args:
			proximity_readings (list): A list containing readings for each proximity sensor.
									Each sensor's readings are stored as a list.
			forward_distance_readings (list): A list containing multiple single-element readings
											for the forward distance sensor.

		Returns:
			tuple: A tuple containing two elements:
				- A list of median readings for each proximity sensor.
				- The median reading for the forward distance sensor.
		"""
	
		# Calculate the median proximity reading for each sensor
		proximity_median = []
		for sensor_readings in proximity_readings:
			median_reading = sorted(sensor_readings)[len(sensor_readings) // 2]  # Find the median value
			proximity_median.append(median_reading)
				
		return proximity_median        
		
	def store_sensor_data(self, proximity_sensors: List[int], forward_distance: float):
		"""Stores sensor data.

		Args:
			proximity_sensors (list of int): List of proximity sensor readings.
			forward_distance (float): Forward distance reading.
		"""
		self.unsafe_sensors.clear()
		filtered_sensor_readings = self.exponential_moving_average_filter(proximity_sensors)
		self.proximity_sensors = filtered_sensor_readings
		self.forward_distance = forward_distance
	
	def calculate_actual_distance(self,proximity_reading):
		"""
		Calculate the actual distance corresponding to a given proximity reading.
		
		Args:
			proximity_reading (float): Proximity reading value.
		
		Returns:
			float: Actual distance.
		"""
		
		# Fixed optimal parameters obtained from curve fitting
		a = 1521.7170108943192
		b = -1.256270043608874
		c = -118.51612931657212
			
		# Calculate actual distance using the inverse of the fitted function
		actual_distance_cm = ((proximity_reading - c) / a) ** (1 / b)
		
		actual_distance_m = round(actual_distance_cm / 100 , 3)     
		
		return actual_distance_m   
   
	def calculate_actual_forward_distance(self, reading):
		"""
		Calculate the actual distance based on the fixed slope and intercept.

		Args:
			reading (float): The reading obtained from the sensor.

		Returns:
			float: The actual distance calculated from the reading.
		"""
		# Fixed slope and intercept
		slope = 0.997
		intercept = -2.686

		actual_distance = slope * reading + intercept
		return round(actual_distance/100, 2)

	def convert_sensor_data_to_actual_distances(self):
		"""
		Convert filtered sensor data to actual distances.
		"""
		# Convert proximity sensor readings to actual distances
		self.actual_proximity_distances = [self.calculate_actual_distance(sensor_readings) for sensor_readings in self.proximity_sensors]
		
		# Convert forward distance reading to actual distance
		self.actual_forward_distance = self.calculate_actual_forward_distance(self.forward_distance) 

	def store_unsafe_sensors(self):
		"""Identify and store sensors with readings less than the safety distance
		"""
		self.unsafe_sensors.clear()                  
		# Iterate through sensor readings
		for sensor_number, distance in enumerate(self.actual_proximity_distances):
			if distance < self.safety_distance:
				self.unsafe_sensors.append((sensor_number, distance)) # Stores sensor number and distance to object
				
	def process_sensors_oa_mode(self, proximity_sensors: List[int], forward_distance: float):
		"""Store all sensor data and separately store the unsafe sensors

		Args:
			proximity_sensors (List[int]): EPUCK sensor readings
			forward_distance (float): distance from proximity sensor
		"""
		self.store_sensor_data(proximity_sensors,forward_distance)
		self.convert_sensor_data_to_actual_distances()
		self.store_unsafe_sensors()
	
	def is_object_on_path(self):
		"""Checks if an object is on the path of the robot

		Returns:
			Boolean: True, if in the valid detection zone. Otherwise, False
		"""
		left_sensor_reading = self.actual_proximity_distances[7]
		right_sensor_reading = self.actual_proximity_distances[0]
		
		if left_sensor_reading > self.safety_distance and right_sensor_reading > self.safety_distance and self.actual_forward_distance > self.safety_distance:
			return False
		else:
			#print("Object ON Path")
			return True         
		
	def calculate_side_variables(self):
		"""Calculate the variables required to determine the side value of the object relative to the robot.
		Variables: Ll and Lr
		"""  
									 
		if not self.is_object_on_path():
			# The outputs of the obstacle avoidance controller are equal to zero when the obstacle does not lie on the robot’s course.
			self.l_l = 0
			self.l_r = 0
		else:
			left_sensor_reading = self.actual_proximity_distances[7]
			right_sensor_reading = self.actual_proximity_distances[0]
			#print("Left sensor reading: ",left_sensor_reading)
			#print("Right sensor reading: ", right_sensor_reading)
			
			total_distance = left_sensor_reading + right_sensor_reading
			if total_distance > 0:
				# When the left sensor reading is greater, the object is to the right of the robot
				# So, LR should be greater than LL
				if left_sensor_reading > right_sensor_reading:
					self.l_l = right_sensor_reading / total_distance
					self.l_r = left_sensor_reading / total_distance
				else:
					# When the right sensor reading is greater, the object is to the left of the robot
					# So, LL should be greater than LR
					self.l_l = right_sensor_reading / total_distance
					self.l_r = left_sensor_reading / total_distance
			else:
				# Case when both sensor readings are at 0 distance / touching obstacle
				print("0 distance case")
				self.l_l = 0.5
				self.l_r = 0.5    
		"""
		print("Left sensor reading: ", left_sensor_reading)
		print("Right sensor reading: ", right_sensor_reading)
		print("total distance:", total_distance)       
		print("Ll: ", round(self.l_l,3))
		print("Lr: ", round(self.l_r,3))
		"""    
	   
	#************************************** Membership functions and fuzzification of the side linguistic variable ***********************************************#
	
	def num_to_range(self,num, inMin, inMax, outMin, outMax):
		"""Linearly maps a value from an input range to an output range.
	
		Args:
			num (float): The value to be mapped.
			inMin (float): The minimum value of the input range.
			inMax (float): The maximum value of the input range.
			outMin (float): The minimum value of the output range.
			outMax (float): The maximum value of the output range.
		
		Returns:
			float: The mapped value.
		"""
		num_mapped = outMin + (float(num - inMin) / float(inMax - inMin) * (outMax
					- outMin))
		return num_mapped
  
	def membership_function_right(self,x):
		"""
		Membership function for the linguistic variable 'right'.
		
		Args:
			x (float): Input value representing the position on the right side.

		Returns:
			str: Membership degree label ('Low', 'Medium', or 'High').
		"""
	   
	   # Convert to apporopriate membership range
		x_mapped = self.num_to_range(x,0,1,0,0.5)
		
		# Define the input ranges and corresponding membership degrees
		low_range = (0, 0.25)
		medium_range = (0.25, 0.375)
		high_range = (0.375, 1)

		# Determine membership degree based on input value
		if low_range[0] <= x_mapped < low_range[1]:
			return 'Low'
		elif medium_range[0] <= x_mapped < medium_range[1]:
			return 'Medium'
		elif x_mapped >= high_range[0]:
			return 'High'
		else:
			return 'Out of Range'

	def membership_function_left(self,x):
		"""
		Membership function for the linguistic variable 'left'.
		
		Args:
			x (float): Input value representing the position on the left side.

		Returns:
			str: Membership degree label ('Low', 'Medium', or 'High').
		"""
	   
		# Convert to apporopriate membership range
		x_mapped = self.num_to_range(x,0,1,0.5,1)
	   
		# Define the input ranges and corresponding membership degrees
		low_range = (0, 0.75)
		medium_range = (0.75, 0.875)
		high_range = (0.875, 1)

		# Determine membership degree based on input value
		if low_range[0] <= x_mapped < low_range[1]:
			return 'Low'
		elif medium_range[0] <= x_mapped < medium_range[1]:
			return 'Medium'
		elif x_mapped >= high_range[0]:
			return 'High'
		else:
			return 'Out of Range'
	
	def fuzzify_side(self):
		"""Fuzzify the side value using membership functions.
		The degree of memebrship for the 'left' and 'right' are stored
		"""
		if not self.is_object_on_path():
			# Changed this from low -> out of range (Does not still seem to print these)
			print("UNDETERMINED: Object NOT on Path")
			left_membership = 'Low'
			right_membership = 'Low'
		else:
			# Object side is determined by the opposite sensor
			left_membership = self.membership_function_left(self.l_l)
			right_membership = self.membership_function_right(self.l_r)
		
		self.left_membership = left_membership
		self.right_membership = right_membership
	
	def determine_side(self):
		"""
		Determine if the object is to 'left' or 'right' relative to the robot
		"""
		side = 'Undetermined'
		if self.left_membership == 'High':
			side ='Left'
		elif self.right_membership == 'High':
			side = 'Right'
		elif self.left_membership == 'Medium' and self.right_membership == 'Medium':
			side = 'Centre'
		elif self.left_membership == 'Medium' and self.right_membership == 'Low':
			side = 'Left'
		elif self.left_membership == 'Low' and self.right_membership == 'Medium':
			side ='Right'
		elif self.left_membership == 'Out of Range' and self.right_membership == 'Out of Range':
			side = 'Out of Path'
		else:
			# Unexpected Cases
			side = 'Undetermined'
		
		self.side = side
	
	def process_side(self):
		"""Fuzzification process of the linguistic variable, 'side'.
		"""
		self.calculate_side_variables()
		self.fuzzify_side()
		self.determine_side()
   
   #********************************** Membership functions and fuzzification of the linguistic variable: distance to object ****************************************#
   
	def membership_function_far(self,x):
		"""
		Membership function quantifying how 'far' for the linguistic variable 'distance to the object'.
		
		Args:
			x (float): Input value representing the distance to the object.

		Returns:
			str: Membership degree label ('High', 'Medium', or 'Low').
		"""

		# Determine membership degree based on input value
		if 0 <= x < 0.0254 or 0.0634 <= x < 0.076:
			return 'Low'
		elif 0.0254 <= x < 0.0317 or 0.507 <= x < 0.0634:
			return 'Medium'
		elif 0.0317 <= x < 0.038 or 0.038 <= x < 0.0507:
			return 'High'
		else:
			return 'Out of Range'

	def membership_function_near(self,x):
		"""
		Membership function quantifying how 'near' for the linguistic variable 'distance to the object'.
		
		Args:
			x (float): Input value representing the distance to the object.

		Returns:
			str: Membership degree label ('High', 'Medium', or 'Low').
		"""
		# Ranges and corresponding membership degrees
		low_range = (0.0317, 0.076)  # Low range covers remaining detection range  
		medium_range = (0.0253, 0.0317)
		high_range = (0,0.0253) 

		# Determine membership degree based on input value
		if low_range[0] <= x < low_range[1]:
			return 'Low'
		elif medium_range[0] <= x < medium_range[1]:
			return 'Medium'
		elif x >= high_range[0]:
			return 'High'
		else:
			return 'Out of Range'

	def fuzzify_distance_to_object(self):
		""" Fuzzify the distance to object value using membership functions.
			"""
		# Sensors at the front of the EPUCK
		left_forward_reading = self.actual_proximity_distances[7]
		right_forward_reading = self.actual_proximity_distances[0]
		centre_forward_reading = self.actual_forward_distance   
		
		# Storing the distance of the nearest object at the front
		shortest_distance = min(centre_forward_reading,left_forward_reading,right_forward_reading)     
	   
		self.near_membership = self.membership_function_near(shortest_distance)
		self.far_membership = self.membership_function_far(shortest_distance)
		
	def process_linguistic_variables(self):
		"""Determines the membership of the linguistic variables (Side and dor)
		"""
		self.process_side()
		self.fuzzify_distance_to_object()
	
	def determine_d_or(self):
		"""
		Determine if the object is 'far' or 'near' based on the comparison of membership degrees.

		Returns:
			str: 'far' or 'near' based on the comparison of membership degrees.
		"""
		if self.far_membership == 'High' or self.near_membership == 'Low':
			return 'far'
		elif self.near_membership == 'High' or self.far_membership == 'Low':
			return 'near'
		elif self.near_membership == 'Medium' and self.far_membership == 'Medium':
			# If both are 'Medium', assuming worst case scenario -> near
			return 'near'
		elif self.near_membership == 'Low' and self.far_membership == 'Low':
			return 'far'
		else:
			# If both memberships are 'Out of Range'
			return 'undetermined'
	
	def apply_fuzzy_rules(self):
		"""Sets the linear and angular velocity based on the linguistic variables: 'side' and 'dor' (distance to object)
		"""
		side = self.side
		d_or = self.determine_d_or()
		
		if side == 'Left' and d_or == 'near':
			self.v_oa = self.v_max/2
			self.w_oa = -(self.w_max)
		elif side == 'Left' and d_or == 'far':
			self.v_oa = self.v_max
			self.w_oa = -(self.w_max)/2
		elif side == 'Right' and d_or == 'near':
			self.v_oa = self.v_max/2
			self.w_oa = self.w_max
		elif side == 'Right' and d_or == 'far':
			self.v_oa = self.v_max
			self.w_oa = self.w_max/2
		else:
			# Object is not on the robot's path
			self.v_oa = 0
			self.w_oa = 0
	
	def process_fuzzification(self):
		"""Applies altered Mitrovic and Djurovic fuzzy-based controller for Differential
			Drive Mobile Robot Obstacle Avoidance.
			
			Processes the linguistic variables to quantify/describe the incoming object's position relative to the robot
		"""
		self.process_linguistic_variables()
		self.apply_fuzzy_rules()
		
	def determine_wheel_velocities(self):
		"""Determines between the choice of velocity for the robot.
		
		Choices: User generated velocity or OA algorithm velocity
		"""
		# Determining Linear Velocity
		if self.v_oa == 0:
			self.v = self.v_u
		else:
			self.v = self.v_oa
			
		# Determining angular velocity
		if self.w_oa == 0:
			self.w = self.w_u
		else:
			self.w = self.w_oa        
				
	def calculate_wheel_velocities(self):
		"""Calculate the speeds of the left and right wheels based on linear and angular velocities.
		"""
		linear_velocity = self.v
		angular_velocity = self.w
		
		self.v_lw_mps = linear_velocity - (self.wheel_distance_m/ 2) * angular_velocity
		self.v_rw_mps = linear_velocity + ( self.wheel_distance_m/ 2) * angular_velocity

	def process_velcoities(self):
		"""Utilises the linguistic variables and the FLC rules to calculate necessary wheel velocities. 
		"""
		self.determine_wheel_velocities()
		self.calculate_wheel_velocities()   
		self.set_velocity(self.v_lw_mps,self.v_rw_mps) 
	
	def run_object_avoidance(self,proximity_sensor_readings,forward_distance_cm):
		""" Processes the necessary functions for object avoidance
		"""
		# set_vibration()
		self.process_sensors_oa_mode(proximity_sensor_readings,forward_distance_cm)
		self.set_led_colour('purple')
		self.process_fuzzification()
		self.process_velcoities()
		
	def is_object_cleared(self):
		"""
		Checks if the obstacle has been cleared based on the 'side' variables: Ll and Lr.

		Returns:
			bool: True if the obstacle is cleared, False otherwise.
		"""
				
		if self.v_lw_mps == 0 and self.v_rw_mps == 0:
		# Object is no longer on the robot's path            
			return True
		else:
			return False
	
	def is_motion_possible(self):
		"""
		Check if motion is possible based on proximity readings from different sensors.

		Returns:
			bool: True if motion is possible, False otherwise.
		"""
		# Define safety distances for sensors
		safety_distance = self.safety_distance

		# Check if back sensors detect obstacles
		if (self.actual_proximity_distances[4] < safety_distance and
				self.actual_proximity_distances[3] < safety_distance):
			# Check if left and right sensors detect obstacles
			if (self.actual_proximity_distances[5] < safety_distance and
					self.actual_proximity_distances[2] < safety_distance):
				# Check if centre left and centre right sensors detect obstacles
				if ((self.actual_proximity_distances[6] < safety_distance or
					self.actual_proximity_distances[7] < safety_distance) and
						(self.actual_proximity_distances[0] < safety_distance or
						self.actual_proximity_distances[1] < safety_distance)):
					return True
		return False

	def alert_static_operator(self):
		"""Operator is alerted.
		No Action is permitted by the robot.
		"""
		self.set_led_colour('red')
	
	# Just using this to mimic the fsm. when using the fsm. colour is set in respective states
	def check_robot_state(self):
		if self.is_object_cleared():
			print("Object Cleared!")
			print("User Controlling mode")
			self.set_led_colour("green")   
		else:
			print("Object not cleared")
	
	def fsm_OA(self, proximity_readings, forward_distance_cm,next_state_main):
		"""
		Determines the robot's behavior based on obstacle detection, then executes corresponding actions.

		Processes sensors to detect obstacles. Sets the robot's state to USER_CONTROLLING if no obstacle 
		is detected, AVOIDING_OBSTACLE if avoidance is possible, or STATIC_OPERATOR_REQUIRED if intervention
		is needed. Executes actions accordingly.
		"""
		self.next_state = next_state_main
		
		# DETERMINE THE STATES
		if self.is_object_cleared():
			# No obstacle in the robot's path and previous obstacle has been cleared
			self.curr_state = USER_CONTROLLING
		else:       
			# Obstacle in the robot's path            
			self.curr_state = AVOIDING_OBSTACLE

		# FSM Execution        
		if self.next_state == USER_CONTROLLING:
			self.run_user_controlling(proximity_readings, forward_distance_cm)             
			if not self.is_forward_allowed:
				self.next_state = AVOIDING_OBSTACLE    
			
		elif self.next_state == AVOIDING_OBSTACLE:
			self.run_object_avoidance(proximity_readings, forward_distance_cm)
			if self.is_object_cleared():
				self.next_state = USER_CONTROLLING
					
		 
	#******************************************************************** TESTING FUNCTIONS *****************************************************************************#
	def print_sensor_data(self, proximity_distances, forward_distance):
		"""
		Print sensor data, replacing the maximum value with 'X' if found.
		
		Args:
			distances (list): List of sensor distances.
		"""
		# Create a new list to store the modified sensor data
		modified_sensor_data = []
		
		# Iterate over each element in distances
		for distance in proximity_distances:
			if distance == self.max_distance_measured:
				modified_sensor_data.append("X")  # Replace max value with 'x/-'
			else:
				modified_sensor_data.append(distance)  # Otherwise, keep the original value
		
		# Print the modified sensor data
		print("Sensor data: ", modified_sensor_data)
		print("Forward distance: ", forward_distance)
	
	def test_stored_sensor_readings(self):
		"""Prints the stored sensor readings."""
		for i, reading in enumerate(self.proximity_sensors):
			print(f"Sensor {i} Reading:", reading)
			
	def test_front_sensor_readings(self,proximity_readings, forward_distance_readings):
		"""Prints the readings of the left and right proximity sensors"""
		
		proximity_ema = self.moving_average_filter(proximity_readings)
		print("Right Reading:", proximity_ema[0])
		print("Left Reading: ", proximity_ema[7])
		
	def test_side_linguistic_variables(self, proximity_sensors, forward_distance):
		"""
		Test function to evaluate the side linguistic variable and its accompanying Ll and Lr values.

		Args:
			proximity_sensors (list): List of proximity sensor readings.
			forward_distance (float): Forward distance reading.
		"""
		
		print("TESTING: 'SIDE' LINGUISTIC VARIABLES")  
		print()
		
		# Process sensor readings and calculate side linguistic variable
		self.process_sensors(proximity_sensors, forward_distance)
		self.process_side()     
		 
		data = {
			"Variable": ["Ll", "Lr", "Left Side Membership", "Right Side Membership", "Side"],
			"Value": [self.l_l, self.l_r, self.left_membership, self.right_membership, self.side]
		}
		
		df = pd.DataFrame(data)
		
		print("Test Results:")
		print(df)
		
	def test_nearness_linguistic_variables(self, proximity_sensors, forward_distance):
		
		"""
		Test function to evaluate the Nearness linguistic variable and its accompanying membership values.

		Args:
			proximity_sensors (list): List of proximity sensor readings.
			forward_distance (float): Forward distance reading.
		"""
		# Process sensor readings and calculate side linguistic variable
		self.process_sensors(proximity_sensors, forward_distance)
		self.fuzzify_distance_to_object()
		d_or = self.determine_d_or()
				
		data = {
			"Variable": ["Near Membership", "Far Membership", "Dor"],
			"Value": [self.near_membership, self.far_membership, d_or]
		}   
				
		df = pd.DataFrame(data)
		
		print("Test Results:")
		print(df)
		print()
  
	def print_object_avoidance_variables(self):
		dor = self.determine_d_or()
		print("Side: ", self.side)
		print("Dor: ", dor)
		
		print()
		
		print("Left wheel speed: ", self.v_lw_mps)
		print("Right wheel speed: ", self.v_rw_mps)

# ====================================== OFFICIAL TESTS ========================================== #

	def test_desired_max_velocity(self, velocity_mps, proximity_readings, forward_distance_cm, file_path, file_name):
		# TESTING COLOUR
		self.set_led_colour("blue")
		
		self.set_velocity(velocity_mps, velocity_mps)
		
		# Testing forward motion only - Due to lack of sensor in the centre back of the epuck
		velocity_mps = abs(velocity_mps)
		
		# Measure distances
		self.process_sensors_user_mode(proximity_readings, forward_distance_cm)
						
		# If distance is less than safety distance
		if self.actual_forward_distance <= self.safety_distance:
			self.is_test_done = True  # Conclude test
			self.set_velocity(0, 0)  # Stop robot

		# Measure distances
		self.process_sensors_user_mode(proximity_readings, forward_distance_cm)
		
		# Print and store results once test is concluded
		if self.is_test_done:
			# Record distance when robot is stopped
			measured_distance = self.actual_forward_distance
			
			# Printing the results
			print()
			print("TEST RESULTS: ")
			print()
			print("Velocity: ", velocity_mps)
			print("Safety distance: ", self.safety_distance)
			print("Measured distance: ", measured_distance)
			
			# Save outcome to Excel file
			self.save_desired_velocity_to_excel(velocity_mps, measured_distance, file_path, file_name)

	def save_desired_velocity_to_excel(self, velocity_mps, measured_distance, file_path, file_name):
		# Combine file path and file name
		full_file_path = os.path.join(file_path, file_name)
		
		# Open or create the Excel file
		try:
			wb = openpyxl.load_workbook(full_file_path)
		except FileNotFoundError:
			wb = openpyxl.Workbook()
		
		sheet = wb.active
		
		# Find the next available row
		next_row = sheet.max_row + 1
		
		# Check if velocity heading is already added
		if sheet.cell(row=1, column=1).value != "Velocity [mps]":
			# Write "Velocity" heading
			sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
			sheet.cell(row=1, column=1, value="Velocity[mps]")
		
		# Check if measured distance headings are already added
		if sheet.cell(row=1, column=2).value != "Measured Distance[m]":
			# Write "Measured Distance" heading
			sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
			sheet.cell(row=1, column=2, value="Measured Distance[m]")
			# Write subheadings for measured distance
			for i, subheading in enumerate(["1", "2", "3", "4", "5"]):
				sheet.cell(row=2, column=i+2, value=subheading)
		
		# Check if the desired speed already exists in the file
		existing_speed_row = None
		for row in range(1, next_row):
			if sheet.cell(row=row, column=1).value == velocity_mps:
				existing_speed_row = row
				break
		
		if existing_speed_row is not None:
			# If the desired speed already exists, use the existing row
			row_to_use = existing_speed_row
		else:
			# If the desired speed is new, use the next available row
			row_to_use = next_row
			# Write the velocity to the first column of the row
			sheet.cell(row=row_to_use, column=1, value=velocity_mps)        
		
		# Find the next available column for the current velocity row
		next_col = 2  # Start from the second column
		while sheet.cell(row=row_to_use, column=next_col).value is not None:
			next_col += 1
		
		# Write the reading to the next column
		sheet.cell(row=row_to_use, column=next_col, value=measured_distance)
		
		# Save the changes to the Excel file
		try:
			wb.save(full_file_path)
			print("Test saved successfully.")
		except Exception as e:
			# Handle save error
			print("An error occurred while saving the Excel file:", e)

	def test_filters(self, actual_distance_cm, sensor_choice, proximity_readings, forward_distance_cm, file_path, file_name):
		"""
		Test various filters with proximity readings and save the results to an Excel file.

		Summary:
		Calculates filtered readings using moving average, exponential moving average, and median filters. Prints the input data,
		processes the sensor data to actual distances, prints the filtered results, and saves them to an Excel file.

		Parameters:
		actual_distance_cm (float): The actual distance.
		sensor_choice (int): The chosen sensor index.
		proximity_readings (list): List of proximity readings.
		forward_distance_cm (float): Forward distance in centimeters.
		file_path (str): Path to the Excel file.
		file_name (str): Name of the Excel file.
		"""

		# TESTING COLOUR
		self.set_led_colour("blue")
	
		# Validate sensor choice
		if sensor_choice not in [0, 1, 2, 3, 4, 5, 6, 7]:
			print("Error: Sensor choice must be an integer between 0 and 7.")
			self.is_test_done = True
			return

		# Validate actual distance
		if actual_distance_cm not in [1, 2, 3, 4, 5, 6]:
			print("Error: Actual distance must be 1, 2, 3, 4, 5, or 6 centimeters.")
			self.is_test_done = True
			return
		
		# Collect sensor data
		self.process_sensors_oa_mode(proximity_readings, forward_distance_cm)        
		
  		# Calculate filtered readings from each filter
		proximity_avg = self.moving_average_filter(proximity_readings)
		proximity_ema = self.exponential_moving_average_filter(proximity_readings)
		proximity_median = self.median_filter(proximity_readings)

		sensor_choice_idx = int(sensor_choice)     # Chosen Sensor
	
		# Convert proximity reading for selecting sensor to distance readings
		distance_avg = self.calculate_actual_distance(proximity_avg[sensor_choice_idx])
		distance_ema = self.calculate_actual_distance(proximity_ema[sensor_choice_idx])
		distance_median = self.calculate_actual_distance(proximity_median[sensor_choice_idx])
		
		# convert actual distance: cm -> m
		actual_distance_m = actual_distance_cm/100

		# Print the input data
		print()
		print("INPUT DATA:")
		print()
		print("Proximity Readings Input:", proximity_readings)
		print("Sensor Choice:", sensor_choice)
		print("Actual Distance:", actual_distance_m,"[m]")
		print()        
		
		# Print the results
		print("RESULTS:")
		print()
		self.print_filter_results(distance_avg, "Moving Average")
		self.print_filter_results(distance_ema, "Exponential Moving Average")
		self.print_filter_results(distance_median, "Median")
		
		# Combine file path and file name
		full_file_path = os.path.join(file_path, file_name)	
		
		# Save the filtered readings to the Excel file
		self.save_filtered_readings_to_excel( actual_distance_m, distance_avg, full_file_path,  "Moving Average")
		self.save_filtered_readings_to_excel( actual_distance_m, distance_ema, full_file_path, "Exponential Moving Average")
		self.save_filtered_readings_to_excel( actual_distance_m, distance_median, full_file_path, "Median")

		print("Test complete.")
		self.is_test_done = True
   
	def print_filter_results(self, filtered_reading, filter_type):
		"""
		Print the filtered reading and actual distance for a given filter type.

		Summary:
		Prints the filtered reading and the actual distance for a specific filter type and sensor choice index.

		Parameters:
		filtered_reading (float): The filtered reading.
		filter_type (str): The type of filter.
		sensor_choice_idx (int): The index of the chosen sensor.
		"""
		
		print(filter_type, "Filter")
		print("Filtered Reading:", filtered_reading)
		print()
	
	def save_filtered_readings_to_excel(self, actual_distance_m, filtered_reading, full_file_path, filter_type ):
		"""
		Save the filtered readings to the Excel file.

		Summary:
		Saves the filtered readings to the specified sheet in the Excel file along with the actual distance.

		Parameters:
		sheet (openpyxl.worksheet.Worksheet): The Excel sheet to write the data to.
		next_row (int): The next available row in the sheet.
		actual_distance (float): The actual distance in m.
		filtered_reading (float): The filtered reading.
		filter_type (str): The type of filter.
		"""
		# Open or create the Excel file
		try:
			wb = openpyxl.load_workbook(full_file_path)
		except FileNotFoundError:
			wb = openpyxl.Workbook()
		
		sheet = wb.active
		
		# Find the next available row
		next_row = sheet.max_row + 1
		# Find the row for the filter
		filter_row = None
		for row in range(1, next_row):
			if sheet.cell(row=row, column=1).value == filter_type:
				filter_row = row
				break
	
		# Find the column for the actual distance
		actual_distance_col = None
		for col in range(2, sheet.max_column + 1):
			if sheet.cell(row=2, column=col).value == actual_distance_m:
				actual_distance_col = col
				break

		# Find the next available reading column under the actual distance column
		next_reading_col = None
		for col in range(actual_distance_col,actual_distance_col + 5 ):
			if sheet.cell(row=filter_row, column=col).value is None:
				next_reading_col = col
				break

		if next_reading_col is None:
			print("Error: Maximum number of readings (5) for this distance already reached.")
			return

		# Fill in the next available column with the filtered reading
		filtered_reading_col = next_reading_col
		sheet.cell(row=filter_row, column=filtered_reading_col, value=filtered_reading)
				
		# Save the changes to the Excel file
		try:
			wb.save(full_file_path)
			print("Reading saved")

		except Exception as e:
			# Handle save error
			print("An error occurred while saving the Excel file:", e)

	def test_width(self, proximity_readings, forward_distance_cm, obstacle_width_cm):
		"""
		Testing the success of the Object Avoidance algorithm with varying obstacle widths

		Args:
			proximity_readings (_type_): List of proximity readings
			forward_distance_cm (_type_): Forward distance in cm
			obstacle_width_cm (_type_): Obstacle width in cm
			file_path (str): Path to the Excel file.
			file_name (str): Name of the Excel file.
		"""
  
		# Validate actual distance
		if obstacle_width_cm not in [1.5, 3, 7, 14, 21, 35]:
			print("Error: Obstacle width must be: 1.5, 3, 7, 14, 21, 35 cm")
			self.is_test_done = True
			return

		obstacle_width_m = obstacle_width_cm / 100
		print()
		print("INPUT DATA:")
		print()
		print("Obstacle width: ",obstacle_width_m)
  
		# Run OA algorithm
		self.run_object_avoidance(proximity_readings, forward_distance_cm)
  
		if not self.is_object_on_path():
			self.is_test_done = True	
			print("Test complete.")

	def save_outcome_width_to_excel(self, obstacle_width_cm, file_path, file_name):
		"""
		Save the outcome of the object avoidance test based on obstacle width to an Excel file.

		Summary:
		Saves the outcome (success or failure) of the object avoidance test based on obstacle width
		to the specified Excel file.

		Parameters:
		- obstacle_width_cm (float): The width of the obstacle in centimeters.
		- file_path (str): The file path of the Excel file directory.
		- file_name (str): The name of the Excel file.
		"""

		# Convert width: cm -> m
		obstacle_width_m = obstacle_width_cm / 100
  
		# Ask user for collision input
		collision_input = input("Was there a collision? (Y/N): ")

		# Set the collision status
		outcome = 'Fail' if collision_input.upper() == 'Y' else 'Success'
  
		print("Recorded value: ", outcome)

		# Combine file path and file name
		full_file_path = os.path.join(file_path, file_name)	
		
  		# Open or create the Excel file
		try:
			wb = openpyxl.load_workbook(full_file_path)
		except FileNotFoundError:
			wb = openpyxl.Workbook()
		
		sheet = wb.active
		
		# Find the next available row
		next_row = sheet.max_row + 1
  		
		# Find the row for the width
		width_row = None
		for row in range(1, next_row):
			if sheet.cell(row=row, column=1).value == obstacle_width_m:
				width_row = row
				break

		# Find the next available reading column under the actual distance column
		next_reading_col = None
		for col in range(2,6):
			if sheet.cell(row=width_row, column=col).value is None:
				next_reading_col = col
				break

		if next_reading_col is None:
			print("Error: Maximum number of readings (5) for this width already reached.")
			return

		# Fill in the next available column with the filtered reading
		outcome_col = next_reading_col
		sheet.cell(row=width_row, column=outcome_col, value=outcome)
				
		# Save the changes to the Excel file
		try:
			wb.save(full_file_path)
			print("Reading saved")

		except Exception as e:
			# Handle save error
			print("An error occurred while saving the Excel file:", e)
   
	def test_shape(self, proximity_readings, forward_distance_cm, obstacle_shape):
		"""
		Testing the success of the Object Avoidance algorithm with varying obstacle shapes

		Args:
			proximity_readings (_type_): List of proximity readings
			forward_distance_cm (_type_): Forward distance in cm
			obstacle_shape (_type_): Obstacle shape description
			file_path (str): Path to the Excel file.
			file_name (str): Name of the Excel file.
		"""
  
		# Validate actual distance
		if obstacle_shape not in ["Uniform (Straight)", "Uniform (Curved)", "Non-uniform"]:
			print("Error: Obstacle shape must be: Uniform (Straight)', 'Uniform (Curved)', 'Non-uniform'")
			self.is_test_done = True
			return

		print()
		print("INPUT DATA:")
		print("Obstacle shape: ",obstacle_shape)
		print()
  
		# Run OA algorithm
		self.run_object_avoidance(proximity_readings, forward_distance_cm)

		if not self.is_object_on_path():
			self.is_test_done = True	
			print("Test complete.")

	def save_outcome_shape_to_excel(self, obstacle_shape, file_path,file_name ):
		"""
		Save the outcome of the object avoidance test based on obstacle shape to an Excel file.

		Summary:
		Saves the outcome (success or failure) of the object avoidance test based on obstacle shape
		to the specified Excel file.

		Parameters:
		- obstacle_shape (str): The shape of the obstacle (e.g., square, circle).
		- file_path (str): The file path of the Excel file directory.
		- file_name (str): The name of the Excel file.
		"""	
  
		# Combine file path and file name
		full_file_path = os.path.join(file_path, file_name)	
		
  		# Ask user for collision input
		collision_input = input("Was there a collision? (Y/N): ")

		# Set the collision status
		outcome = 'Fail' if collision_input.upper() == 'Y' else 'Success'
  
		print("Recorded value: ", outcome)
  
		# Open or create the Excel file
		try:
			wb = openpyxl.load_workbook(full_file_path)
		except FileNotFoundError:
			wb = openpyxl.Workbook()
		
		sheet = wb.active
		
		# Find the next available row
		next_row = sheet.max_row + 1
  
  		# Find the row for the length
		length_row = None
		for row in range(1, next_row):
			if sheet.cell(row=row, column=1).value == obstacle_shape:
				length_row = row
				break

		# Find the next available reading column under the outcome column
		next_reading_col = None
		for col in range(2,6):
			if sheet.cell(row=length_row, column=col).value is None:
				next_reading_col = col
				break

		if next_reading_col is None:
			print("Error: Maximum number of readings (5) for this shape already reached.")
			return

		# Fill in the next available column with the outcome
		outcome_col = next_reading_col
		sheet.cell(row=length_row, column=outcome_col, value=outcome)
				
		# Save the changes to the Excel file
		try:
			wb.save(full_file_path)
			print("Reading saved")

		except Exception as e:
			# Handle save error
			print("An error occurred while saving the Excel file:", e)
   
	def test_OA(self, proximity_readings, forward_distance_cm, object_side, object_nearness):
		"""
		Testing the success of the Object Avoidance algorithm with varying obstacle lengths

		Args:
			proximity_readings (List): List of proximity readings
			forward_distance_cm (float): Forward distance in cm
			object_side (str): Side of object relative to the robot
			object_nearness (str): Nearness of obstacle to robot
			file_path (str): Path to the Excel file.
			file_name (str): Name of the Excel file.
		"""
  
		# Validate object side
		if object_side not in ["Left", "Right", "Centre" , "Out of Path"]:
			print("Error: Obstacle side must be: 'Left', 'Right', 'Centre' or 'Out of Path'")
			self.is_test_done = True
			return

		print()
		print("INPUT DATA:")
		print()
		print("Side: ",object_side)
		print("Nearness: ",object_nearness)
  
		# Run OA algorithm
		self.run_object_avoidance(proximity_readings, forward_distance_cm)

		if not self.is_object_on_path():
			print("Test complete.")
			self.is_test_done = True

	def save_outcome_OA_to_excel(self, object_side, object_nearness, file_path, file_name):
		"""
		Save the outcome of object avoidance to an Excel file.

		Summary:
		Saves the outcome (success or failure) of the object avoidance operation to the specified Excel file,
		organized by the side of the object and its nearness to the robot.

		Parameters:
		- object_side (str): The side of the object (e.g., left, right).
		- object_nearness (str): The nearness of the object to the robot (e.g., close, far).
		- file_path (str): The file path of the Excel file directory.
		- file_name (str): The name of the Excel file.
		"""	
  		# Ask user for collision input
		collision_input = input("Was there a collision? (Y/N): ")

		# Set the collision status
		outcome = 'Fail' if collision_input.upper() == 'Y' else 'Success'
  
		print("Recorded value: ", outcome)

		# Combine file path and file name
		full_file_path = os.path.join(file_path, file_name)			
		
		# Open or create the Excel file
		try:
			wb = openpyxl.load_workbook(full_file_path)
		except FileNotFoundError:
			wb = openpyxl.Workbook()
		
		sheet = wb.active
		
		# Find the next available row
		next_row = sheet.max_row + 1
  
  		# Find the row for the side and nearnes
		side_nearness_row = None
		for row in range(1, next_row):
			if sheet.cell(row=row, column=1).value == object_side and sheet.cell(row=row, column=2).value == object_nearness:
				side_nearness_row = row
				break

		# Find the next available reading column under the outcome column
		next_reading_col = None
		for col in range(3,8):
			if sheet.cell(row=side_nearness_row, column=col).value is None:
				next_reading_col = col
				break

		if next_reading_col is None:
			print("Error: Maximum number of readings (5) for this shape already reached.")
			return

		# Fill in the next available column with the outcome
		outcome_col = next_reading_col
		sheet.cell(row=side_nearness_row, column=outcome_col, value=outcome)
				
		# Save the changes to the Excel file
		try:
			wb.save(full_file_path)
			print("Reading saved")

		except Exception as e:
			# Handle save error
			print("An error occurred while saving the Excel file:", e)
	
	def test_time_efficiency(self, proximity_readings, forward_distance_cm, time_taken, no_readings):
		"""
		Testing the success of the Object Avoidance algorithm with varying obstacle lengths

		Args:
			proximity_readings (List): List of proximity readings
			forward_distance_cm (float): Forward distance in cm
			time_taken (float): Side of object relative to the robot
			no_readings (int): Nearness of obstacle to robot
			file_path (str): Path to the Excel file.
			file_name (str): Name of the Excel file.
		"""

		# Validate sensor choice
		if no_readings not in [1, 2, 3, 4, 5, 6, 7,8,9,10,11,12,13,14,15,16,17,18,19,20]:
			print("Error: Sensor choice must be an integer between 1 and 10.")
			self.is_test_done = True
			return

		print()
		print("INPUT DATA:")
		print()
		print("Time: ",round((time_taken),3 ), " [s]")
  
		# Run OA algorithm
		self.run_object_avoidance(proximity_readings, forward_distance_cm)		

		# Stop executing OA algorithm if obstacle is not on path (while loop in main code)
		if not self.is_object_on_path():
			print("Test complete.")
			self.is_test_done = True
  
	def save_outcome_time_efficiency_to_excel(self, time_taken, no_readings, file_path, file_name):
		"""
		Save the outcome and time efficiency to an Excel file.

		Summary:
		Saves the outcome (success or failure) and the time efficiency (time taken) to the specified Excel file,
		organized by the number of readings.

		Parameters:
		- time_taken (float): The time taken for the operation.
		- no_readings (int): The number of readings.
		- file_path (str): The file path of the Excel file directory.
		- file_name (str): The name of the Excel file.
		"""
		# Combine file path and file name
		full_file_path = os.path.join(file_path, file_name)
  
  		# Ask user for collision input
		collision_input = input("Was there a collision? (Y/N): ")

		# Set the collision status
		outcome = 'Fail' if collision_input.upper() == 'Y' else 'Success'
  
		print("Recorded value: ", outcome)	
  
  		# Open or create the Excel file
		try:
			wb = openpyxl.load_workbook(full_file_path)
		except FileNotFoundError:
			wb = openpyxl.Workbook()
		
		sheet = wb.active
		
		# Find the next available row
		next_row = sheet.max_row + 1
  
  		# Find the row for the number of readings
		side_nearness_row = None
		for row in range(1, next_row):
			if sheet.cell(row=row, column=1).value == no_readings:
				side_nearness_row = row
				break

		# Find the next available reading column under the time taken
		next_reading_time_col = None
		for col in range(2,7):
			if sheet.cell(row=side_nearness_row, column=col).value is None:
				next_reading_time_col = col
				break

		if next_reading_time_col is None:
			print("Error: Maximum number of readings (5) for this reading number reached.")
			return

		# Fill in the next available column with the time taken
		time_col = next_reading_time_col
		sheet.cell(row=side_nearness_row, column=time_col, value=time_taken)
		
  
		# Find the next available reading column under the time taken
		next_reading_outcome_col = next_reading_time_col + 5
		# Fill in the next available column with the time taken
		outcome_col = next_reading_outcome_col
		sheet.cell(row=side_nearness_row, column=outcome_col, value=outcome)
		

		# Save the changes to the Excel file
		try:
			wb.save(full_file_path)
			print("Readings saved")

		except Exception as e:
			# Handle save error
			print("An error occurred while saving the Excel file:", e)